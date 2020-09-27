from tkinter import *   
import openpyxl as xl
import random as rd 

global filename
filename = 'D:\\ApplicationFiles\\FIFATeamGenerator\\files\\Football.xlsx'

def refreshTeam():
    goalKeeper()
    leftBack()
    rightBack()
    centreBack()
    midFielders()
    forward()

def goalKeeper():
        wb = xl.load_workbook(filename)
        sheet = wb['GK']
        row = rd.randrange(2, sheet.max_row + 1)
        player_details = {}

        for col in range(1, sheet.max_column + 1):
            cell_key = sheet.cell(1, col)
            cell_key_value = sheet.cell(row, col)
            player_details[cell_key.value] = cell_key_value.value

        global gk_details_store
        gk_details_store = ''
        for key, value in player_details.items():
            gk_details_store += f'{key} : {value}\n'

        gk_name['text'] = player_details['Jersey Name']

def displayGoalKeeperDetails():
    try:
        full_details['text'] = gk_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def leftBack():
    wb = xl.load_workbook(filename)
    sheet = wb['LB']
    row = rd.randrange(2, sheet.max_row + 1)
    player_details = {}

    for col in range(1, sheet.max_column + 1):
        cell_key = sheet.cell(1, col)
        cell_key_value = sheet.cell(row, col)
        player_details[cell_key.value] = cell_key_value.value

    global lb_details_store
    lb_details_store = ''
    for key, value in player_details.items():
        lb_details_store += f'{key} : {value}\n'

    lb_name['text'] = player_details['Jersey Name']
    return player_details

def displayLeftBackDetails():
    try:
        full_details['text'] = lb_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def rightBack():
    wb = xl.load_workbook(filename)
    sheet = wb['RB']
    row = rd.randrange(2, sheet.max_row + 1)
    player_details = {}

    for col in range(1, sheet.max_column + 1):
        cell_key = sheet.cell(1, col)
        cell_key_value = sheet.cell(row, col)
        player_details[cell_key.value] = cell_key_value.value

    global rb_details_store
    rb_details_store = ''
    for key, value in player_details.items():
        rb_details_store += f'{key} : {value}\n'

    rb_name['text'] = player_details['Jersey Name']
    return player_details

def displayRightBackDetails():
    try:
        full_details['text'] = rb_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def centreBack():
    wb = xl.load_workbook(filename)
    sheet = wb['CB']
    while True:
        row1 = rd.randrange(2, sheet.max_row + 1)
        row2 = rd.randrange(2, sheet.max_row + 1)
        if row1 != row2:
            break
    player_details_1 = {}
    player_details_2 = {}

    for col in range(1, sheet.max_column + 1):
        cell_key = sheet.cell(1, col)
        cell_key_value = sheet.cell(row1, col)
        player_details_1[cell_key.value] = cell_key_value.value

    for col in range(1, sheet.max_column + 1):
        cell_key = sheet.cell(1, col)
        cell_key_value = sheet.cell(row2, col)
        player_details_2[cell_key.value] = cell_key_value.value

    global cb1_details_store, cb2_details_store
    cb1_details_store = cb2_details_store = ''

    for key, value in player_details_1.items():
        cb1_details_store += f'{key} : {value}\n'

    for key, value in player_details_2.items():
        cb2_details_store += f'{key} : {value}\n'

    cb1_name['text'] = player_details_1['Jersey Name']
    cb2_name['text'] = player_details_2['Jersey Name']

def displayCentreBack1Details():
    try:
        full_details['text'] = cb1_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def displayCentreBack2Details():
    try:
        full_details['text'] = cb2_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def midFielders():
    wb = xl.load_workbook(filename)
    sheet = wb['CM']
    while True:
        row1 = rd.randrange(2, sheet.max_row + 1)
        row2 = rd.randrange(2, sheet.max_row + 1)
        row3 = rd.randrange(2, sheet.max_row + 1)
        if row1 != row2 != row3 != row1:
            break
    player_details_1 = {}
    player_details_2 = {}
    player_details_3 = {}
    
    for col in range(1, sheet.max_column + 1):
         
        cell_key = sheet.cell(1, col)

        cell_key_value_1 = sheet.cell(row1, col)
        player_details_1[cell_key.value] = cell_key_value_1.value

        cell_key_value_2 = sheet.cell(row2, col)
        player_details_2[cell_key.value] = cell_key_value_2.value

        cell_key_value_3 = sheet.cell(row3, col)
        player_details_3[cell_key.value] = cell_key_value_3.value

    global cm_details_store, lm_details_store, rm_details_store
    cm_details_store = lm_details_store = rm_details_store = ''

    for key, value in player_details_1.items():
        cm_details_store += f'{key} : {value}\n'
    for key, value in player_details_2.items():
        lm_details_store += f'{key} : {value}\n'
    for key, value in player_details_3.items():
        rm_details_store += f'{key} : {value}\n'

    cm_name['text'] = player_details_1['Jersey Name']
    lm_name['text'] = player_details_2['Jersey Name']
    rm_name['text'] = player_details_3['Jersey Name']

def displayCentreMidDetails():
    try:
        full_details['text'] = cm_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def displayLeftMidDetails():
    try:
        full_details['text'] = lm_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def displayRightMidDetails():
    try:
        full_details['text'] = rm_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def forward():
    wb = xl.load_workbook(filename)
    sheet = wb['F']
    while True:
        row1 = rd.randrange(2, sheet.max_row + 1)
        row2 = rd.randrange(2, sheet.max_row + 1)
        row3 = rd.randrange(2, sheet.max_row + 1)
        if row1 != row2 != row3 != row1:
            break
    player_details_1 = {}
    player_details_2 = {}
    player_details_3 = {}
 
    for col in range(1, 8):
          
        cell_key = sheet.cell(1, col)

        cell_key_value_1 = sheet.cell(row1, col)
        player_details_1[cell_key.value] = cell_key_value_1.value

        cell_key_value_2 = sheet.cell(row2, col)
        player_details_2[cell_key.value] = cell_key_value_2.value

        cell_key_value_3 = sheet.cell(row3, col)
        player_details_3[cell_key.value] = cell_key_value_3.value
    
    global st_details_store, lw_details_store, rw_details_store
    st_details_store = lw_details_store = rw_details_store = ''

    for key, value in player_details_1.items():
        st_details_store += f'{key} : {value}\n'
    for key, value in player_details_2.items():
        lw_details_store += f'{key} : {value}\n'
    for key, value in player_details_3.items():
        rw_details_store += f'{key} : {value}\n'

    st_name['text'] = player_details_1['Jersey Name']
    lw_name['text'] = player_details_2['Jersey Name']
    rw_name['text'] = player_details_3['Jersey Name']

def displayStrikerDetails():
    try:
        full_details['text'] = st_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def displayLeftWingerDetails():
    try:
        full_details['text'] = lw_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

def displayRightWingerDetails():
    try:
        full_details['text'] = rw_details_store
    except NameError:
        full_details['text'] = 'Refresh First!'

if __name__ == '__main__':
    root = Tk() 
    root.geometry('500x650')    
    root.resizable(False, False)  
    root.config(background = '#DEDFEE')  
    root.title('FIFATeamGenerator(v1.0)')  

    img = PhotoImage(file="D:\\ApplicationFiles\\FIFATeamGenerator\\files\\jersey.png")
    
    # GOALKEEPER
    gk = Label(root,
        image = img,
        bg = '#DEDFEE'
        ).place(x = 220, y = 0)

    gk_name = Button(root,
        text = 'GoalKeeper',
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        font = ('Arial Bold', 7),
        width = 10,
        command = displayGoalKeeperDetails 
        )
    gk_name.place(x = 219, y = 70)

    # CENTRE-BACKS
    cb1 = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 140, y = 100)

    cb1_name = Button(root,
        text = 'Centre-back 1',
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        font = ('Arial Bold', 7),
        command = displayCentreBack1Details
        )
    cb1_name.place(x = 139, y = 170)

    cb2 = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 300, y = 100)

    cb2_name = Button(root,
        text = 'Centre-back 2',
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        font = ('Arial Bold', 7),
        command = displayCentreBack2Details
        )
    cb2_name.place(x = 299, y = 170)

    # LEFT BACK
    lb = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 420, y = 120)

    lb_name = Button(root,
        text = 'LeftBack',
        font = ('Arial Bold', 7),
        width = 10,
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        bg = '#FFD833',
        command = displayLeftBackDetails
        )
    lb_name.place(x = 419, y = 190)

    # RIGHT BACK
    rb = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 20, y = 120)

    rb_name = Button(root,
        text = 'RightBack',
        font = ('Arial Bold', 7),
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        command = displayRightBackDetails
        )
    rb_name.place(x = 19, y = 190)

    # CENTRE MIDFIELDER
    cm = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 220, y = 240)

    cm_name = Button(root,
        text = 'CentreMid',
        font = ('Arial Bold', 7),
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        command = displayCentreMidDetails
        )
    cm_name.place(x = 219, y = 310)

    # LEFT MIDFIELDER
    lm = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 360, y = 240)

    lm_name = Button(root,
        text = 'LeftMid',
        font = ('Arial Bold', 7),
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        command = displayLeftMidDetails
        )
    lm_name.place(x = 359, y = 310)

    # RIGHT MIDFIELDER
    rm = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 80, y = 240)

    rm_name = Button(root,
        text = 'RightMid',
        font = ('Arial Bold', 7),
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        command = displayRightMidDetails
        )
    rm_name.place(x = 79, y = 310)

    # STRIKER
    st = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 220, y = 400)

    st_name = Button(root,
        text = 'Striker',
        font = ('Arial Bold', 7),
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        command = displayStrikerDetails
        )
    st_name.place(x = 219, y = 470)

    # LEFT WINGER
    lw = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 360, y = 370)

    lw_name = Button(root,
        text = 'LeftWinger',
        font = ('Arial Bold', 7),
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        command = displayLeftWingerDetails
        )
    lw_name.place(x = 359, y = 440)

    # RIGHT WINGER
    rw = Label(root,
        bg = '#DEDFEE',
        image = img
        ).place(x = 80, y = 370)

    rw_name = Button(root,
        text = 'RightWinger',
        font = ('Arial Bold', 7),
        width = 10,
        bg = '#FFD833',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        command = displayRightWingerDetails
        )
    rw_name.place(x = 79, y = 440)
    
    # GENERATE NEW TEAM
    refresh_btn = Button(root,
        text = 'Refresh',
        width = 70,
        bg = '#FFD833',
        command = refreshTeam,
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        font = ('Arial Bold', 8)
        )
    refresh_btn.place(x = 0, y = 500)

    # DISPLAY PLAYER DETAILS
    full_details = Label(root,
        text = '',
        fg = '#FFFFFF',
        bg = '#000119',
        font = ('Arial Bold', 8),
        relief = RIDGE
        )
    full_details.place(x = 30, y = 530) 
    
    # FORMATION SELECT
    formation = Menubutton(root,
        text = 'Formation',
        bg = '#DEDFEE',
        fg = '#DEDFEE',
        activebackground = '#E3001A',
        activeforeground = '#FFD833',
        relief = RAISED)
    formation.grid()
    formation.menu = Menu(formation, 
        tearoff = 0,
        bg = '#DEDFEE',
        fg = '#E3001A')
    formation['menu'] = formation.menu

    formation.menu.add_command(label = '3-4-3', command = lambda: print('3-4-3'))
    formation.menu.add_command(label = '4-3-3', command = lambda: print('4-3-3'))
    formation.menu.add_command(label = '4-4-2', command = lambda: print('4-4-2'))
    formation.menu.add_command(label = '5-3-2', command = lambda: print('5-3-2'))

    formation.place(x = 400, y = 530)

    root.mainloop() 
